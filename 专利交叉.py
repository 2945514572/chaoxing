import os
from openpyxl import load_workbook


def write_data_to_txt(excel_file, folder_path):
    # 加载Excel文件
    workbook = load_workbook(excel_file, read_only=True)
    sheet = workbook.active  # 假设第一个sheet是要处理的数据

    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始遍历，跳过表头
        stock_code = str(row[0]).zfill(6)  # 获取股票代码并补齐到6位
        year = str(int(row[2]))  # 获取年份并转为整数再转为字符串

        # 构建txt文件名
        file_name_pattern = f"{stock_code}_*_*.txt"
        file_pattern = os.path.join(folder_path, file_name_pattern)

        # 查找匹配的文件
        matched_files = [f for f in os.listdir(folder_path) if
                         os.path.isfile(os.path.join(folder_path, f)) and f.startswith(stock_code)]

        for file in matched_files:
            parts = file.split('_')
            file_year = parts[2].split('年')[0]  # 提取文件名中的年份
            if year == file_year:
                file_path = os.path.join(folder_path, file)
                # 写入数据到txt文件
                with open(file_path, 'a',encoding='utf-8') as txt_file:
                    column4_data = str(row[3])
                    column5_data = str(row[4])
                    txt_file.write(column4_data + '\t' + column5_data + '\n')

    print('数据写入完成！')


# 调用函数进行处理
excel_file = 'D:/tools/Py代码/A股/patent_updated.xlsx'  # Excel文件路径
folder_path = 'D:/年报txt/TXT版'  # 存放txt文件的文件夹路径
print('运行中。。。。。。')
write_data_to_txt(excel_file, folder_path)
